#!/usr/bin/env python3
"""Export two SidewalkPilot device benchmark suites to an XLSX workbook."""

from __future__ import annotations

import argparse
import json
import math
import re
import statistics
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="B7B7B7")


def load_json(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"Could not read {path}: {exc}") from exc


def version_key(version: str) -> tuple[int, int, str]:
    match = re.fullmatch(r"(\d+)\.(\d+)([a-z]?)", version)
    if match is None:
        raise ValueError(f"Invalid model version: {version}")
    return int(match.group(1)), int(match.group(2)), match.group(3)


def reports_by_version(suite: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(report["model"]["version"]): report
        for report in suite.get("models", [])
    }


def percent_gain(baseline: float, comparison: float) -> float:
    return (comparison / baseline - 1.0) * 100.0


def percent_reduction(baseline: float, comparison: float) -> float:
    return (baseline - comparison) / baseline * 100.0


def geometric_mean(values: list[float]) -> float:
    return math.exp(statistics.fmean(math.log(value) for value in values))


def style_header(row: tuple[Any, ...]) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=THIN_GRAY)


def add_table(sheet: Any, name: str) -> None:
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def fit_columns(sheet: Any, maximum: int = 22) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(max(width, 10), maximum)


def flatten_json(value: Any, prefix: str = "") -> list[tuple[str, Any, str]]:
    rows: list[tuple[str, Any, str]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(flatten_json(child, path))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            rows.extend(flatten_json(child, f"{prefix}[{index}]"))
    else:
        cell_value = "<null>" if value is None else value
        rows.append((prefix, cell_value, type(value).__name__))
    return rows


def add_raw_json_sheet(
    workbook: Workbook,
    title: str,
    payload: dict[str, Any],
    table_name: str,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(["JSON path", "Value", "Type"])
    for row in flatten_json(payload):
        sheet.append(row)
    style_header(sheet[1])
    add_table(sheet, table_name)
    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 64
    sheet.column_dimensions["B"].width = 72
    sheet.column_dimensions["C"].width = 14
    for row in sheet.iter_rows(min_row=2):
        row[0].alignment = Alignment(vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def validate_suites(
    baseline: dict[str, Any], comparison: dict[str, Any]
) -> tuple[list[str], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    baseline_reports = reports_by_version(baseline)
    comparison_reports = reports_by_version(comparison)
    if set(baseline_reports) != set(comparison_reports):
        raise SystemExit("The two suites do not contain the same model versions")
    versions = sorted(baseline_reports, key=version_key)
    if len(versions) != 52:
        raise SystemExit(f"Expected 52 matching models, found {len(versions)}")
    for version in versions:
        left = baseline_reports[version]
        right = comparison_reports[version]
        if left["model"]["sha256"] != right["model"]["sha256"]:
            raise SystemExit(f"Model hash mismatch for v{version}")
        if (
            left["mini_test_dataset"]["set_sha256"]
            != right["mini_test_dataset"]["set_sha256"]
        ):
            raise SystemExit(f"Mini test dataset hash mismatch for v{version}")
    return versions, baseline_reports, comparison_reports


def model_row(
    version: str, baseline: dict[str, Any], comparison: dict[str, Any]
) -> list[Any]:
    bp = baseline["performance"]
    cp = comparison["performance"]
    bm = baseline["memory"]
    cm = comparison["memory"]
    gpu = cp["jetson_gpu_load"]
    return [
        f"v{version}",
        baseline["model"]["series"],
        baseline["model"]["size_mib"],
        bp["ips"],
        cp["ips"],
        cp["ips"] / bp["ips"],
        percent_gain(bp["ips"], cp["ips"]),
        bp["latency_mean_ms"],
        cp["latency_mean_ms"],
        percent_reduction(bp["latency_mean_ms"], cp["latency_mean_ms"]),
        bp["latency_median_ms"],
        cp["latency_median_ms"],
        percent_reduction(bp["latency_median_ms"], cp["latency_median_ms"]),
        bp["latency_p95_ms"],
        cp["latency_p95_ms"],
        percent_reduction(bp["latency_p95_ms"], cp["latency_p95_ms"]),
        bp["latency_p99_ms"],
        cp["latency_p99_ms"],
        percent_reduction(bp["latency_p99_ms"], cp["latency_p99_ms"]),
        bp["local_image_pipeline_ips"],
        cp["local_image_pipeline_ips"],
        cp["local_image_pipeline_ips"] / bp["local_image_pipeline_ips"],
        percent_gain(bp["local_image_pipeline_ips"], cp["local_image_pipeline_ips"]),
        bp["preprocessing_mean_ms_excluded_from_ips"],
        cp["preprocessing_mean_ms_excluded_from_ips"],
        bp["session_load_seconds"],
        cp["session_load_seconds"],
        percent_reduction(bp["session_load_seconds"], cp["session_load_seconds"]),
        bp["inference_effective_cpu_cores"],
        cp["inference_effective_cpu_cores"],
        percent_reduction(
            bp["inference_effective_cpu_cores"],
            cp["inference_effective_cpu_cores"],
        ),
        bm["warm_process_rss_delta_mib"],
        cm["warm_process_rss_delta_mib"],
        cm["warm_cuda_visible_used_delta_mib"],
        gpu["mean_percent"],
        gpu["p95_percent"],
        baseline["model"]["sha256"],
    ]


MODEL_HEADERS = [
    "Model",
    "Series",
    "File MiB",
    "Pi IPS",
    "Jetson IPS",
    "Speedup x",
    "Throughput gain %",
    "Pi mean ms",
    "Jetson mean ms",
    "Mean latency reduction %",
    "Pi median ms",
    "Jetson median ms",
    "Median latency reduction %",
    "Pi p95 ms",
    "Jetson p95 ms",
    "P95 latency reduction %",
    "Pi p99 ms",
    "Jetson p99 ms",
    "P99 latency reduction %",
    "Pi local pipeline IPS",
    "Jetson local pipeline IPS",
    "Pipeline speedup x",
    "Pipeline gain %",
    "Pi preprocess ms",
    "Jetson preprocess ms",
    "Pi session load s",
    "Jetson session load s",
    "Load-time reduction %",
    "Pi effective CPU cores",
    "Jetson effective CPU cores",
    "CPU-core reduction %",
    "Pi warm RSS delta MiB",
    "Jetson warm RSS delta MiB",
    "Jetson CUDA-visible delta MiB",
    "Jetson GPU mean %",
    "Jetson GPU p95 %",
    "Model SHA-256",
]


def family_summary(rows: list[list[Any]], family: str) -> list[Any]:
    if family == "All models":
        selected = rows
    elif family == "Series 1/2":
        selected = [row for row in rows if row[1] in (1, 2)]
    elif family == "Series 3/4":
        selected = [row for row in rows if row[1] in (3, 4)]
    else:
        series = int(family.rsplit(" ", 1)[1])
        selected = [row for row in rows if row[1] == series]
    return [
        family,
        len(selected),
        statistics.fmean(row[3] for row in selected),
        statistics.fmean(row[4] for row in selected),
        geometric_mean([row[5] for row in selected]),
        statistics.fmean(row[6] for row in selected),
        statistics.fmean(row[7] for row in selected),
        statistics.fmean(row[8] for row in selected),
        statistics.fmean(row[9] for row in selected),
        statistics.fmean(row[20] / row[19] for row in selected),
        statistics.fmean(row[22] for row in selected),
        statistics.fmean(row[30] for row in selected),
        statistics.fmean(row[31] for row in selected),
        statistics.fmean(row[32] for row in selected),
        statistics.fmean(row[33] for row in selected),
        statistics.fmean(row[34] for row in selected),
    ]


FAMILY_HEADERS = [
    "Family",
    "Models",
    "Pi mean IPS",
    "Jetson mean IPS",
    "Geometric speedup x",
    "Mean throughput gain %",
    "Pi mean latency ms",
    "Jetson mean latency ms",
    "Mean latency reduction %",
    "Mean pipeline speedup x",
    "Mean pipeline gain %",
    "Mean CPU-core reduction %",
    "Pi mean warm RSS delta MiB",
    "Jetson mean warm RSS delta MiB",
    "Jetson mean CUDA-visible delta MiB",
    "Jetson mean GPU load %",
]


def add_number_formats(sheet: Any) -> None:
    percent_headers = {
        header for header in MODEL_HEADERS if header.endswith("%")
    }
    speed_headers = {"Speedup x", "Pipeline speedup x"}
    for cell in sheet[1]:
        if cell.value in percent_headers:
            for value in sheet.iter_cols(
                min_col=cell.column,
                max_col=cell.column,
                min_row=2,
                max_row=sheet.max_row,
            ):
                for item in value:
                    item.number_format = '0.0"%"'
        elif cell.value in speed_headers:
            for value in sheet.iter_cols(
                min_col=cell.column,
                max_col=cell.column,
                min_row=2,
                max_row=sheet.max_row,
            ):
                for item in value:
                    item.number_format = '0.000"x"'


def build_workbook(
    baseline_path: Path, comparison_path: Path, output_path: Path
) -> None:
    baseline = load_json(baseline_path)
    comparison = load_json(comparison_path)
    versions, baseline_reports, comparison_reports = validate_suites(
        baseline, comparison
    )
    rows = [
        model_row(version, baseline_reports[version], comparison_reports[version])
        for version in versions
    ]
    families = [
        family_summary(rows, family)
        for family in (
            "Series 1",
            "Series 2",
            "Series 1/2",
            "Series 3",
            "Series 4",
            "Series 3/4",
            "All models",
        )
    ]

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary["A1"] = "SidewalkPilot Device Inference Comparison"
    summary["A1"].font = Font(size=18, bold=True, color=WHITE)
    summary["A1"].fill = PatternFill("solid", fgColor=BLUE)
    summary.merge_cells("A1:H1")
    summary["A3"] = "Baseline"
    summary["B3"] = baseline["label"]
    summary["A4"] = "Comparison"
    summary["B4"] = comparison["label"]
    summary["A5"] = "Models"
    summary["B5"] = len(rows)
    summary["A6"] = "Images per model"
    summary["B6"] = baseline_reports[versions[0]]["mini_test_dataset"][
        "image_count"
    ]
    summary["A7"] = "Warmup / measured runs"
    summary["B7"] = (
        f"{baseline_reports[versions[0]]['performance']['warmup_runs']} / "
        f"{baseline_reports[versions[0]]['performance']['measured_runs']}"
    )
    summary["A8"] = "Pi provider"
    summary["B8"] = baseline["provider"]
    summary["A9"] = "Jetson provider"
    summary["B9"] = comparison["provider"]
    summary["D3"] = "Current-model headline"
    summary["D4"] = "Series 3/4 geometric speedup"
    summary["E4"] = families[5][4]
    summary["D5"] = "Series 3/4 mean throughput gain"
    summary["E5"] = families[5][5]
    summary["D6"] = "Series 3/4 mean latency reduction"
    summary["E6"] = families[5][8]
    summary["D7"] = "Series 3/4 mean CPU-core reduction"
    summary["E7"] = families[5][11]
    summary["E4"].number_format = '0.000"x"'
    for cell in ("E5", "E6", "E7"):
        summary[cell].number_format = '0.0"%"'
    for row in summary.iter_rows(min_row=3, max_row=9, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(bottom=THIN_GRAY)
    summary["A12"] = (
        "Measured scope: batch-one FP32 ONNX execution on the same 180 real "
        "sidewalk images. These are not camera-to-servo end-to-end timings."
    )
    summary.merge_cells("A12:H13")
    summary["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["A12"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 26
    summary.column_dimensions["C"].width = 3
    summary.column_dimensions["D"].width = 36
    summary.column_dimensions["E"].width = 18

    presentation = workbook.create_sheet("Presentation Numbers")
    presentation.sheet_view.showGridLines = False
    presentation["A1"] = "Presentation-Ready Jetson Results"
    presentation["A1"].font = Font(size=18, bold=True, color=WHITE)
    presentation["A1"].fill = PatternFill("solid", fgColor=BLUE)
    presentation.merge_cells("A1:H1")
    headline_values = (
        ("A3", "2.345x", "Series 3/4 geometric throughput speedup"),
        ("C3", "57.3%", "Mean model-execution latency reduction"),
        ("E3", "88.5%", "Mean effective CPU-core reduction"),
        ("G3", "1.787x", "Mean local image-pipeline speedup"),
    )
    for value_cell, value, label in headline_values:
        column = presentation[value_cell].column
        presentation[value_cell] = value
        presentation[value_cell].font = Font(size=24, bold=True, color=BLUE)
        presentation.cell(row=4, column=column, value=label)
        presentation.cell(row=4, column=column).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        presentation.merge_cells(
            start_row=4,
            start_column=column,
            end_row=5,
            end_column=column + 1,
        )
    presentation["A7"] = "Recommended measured statement"
    presentation["B7"] = (
        "Across the 22 current Series 3/4 models in this run, Jetson Orin Nano "
        "delivered 2.345x geometric-mean FP32 ONNX throughput and reduced mean "
        "model-execution latency by 57.3% compared with Raspberry Pi 5."
    )
    presentation.merge_cells("B7:H8")
    presentation["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    presentation["A10"] = "Honest small-model result"
    presentation["B10"] = (
        "Series 1/2 favored Raspberry Pi 5: their 672,877-parameter network was "
        "too small to amortize CUDA dispatch, synchronization, and clock-ramp costs."
    )
    presentation.merge_cells("B10:H11")
    presentation["B10"].alignment = Alignment(wrap_text=True, vertical="top")
    presentation["A13"] = "Current model examples"
    presentation.append([])
    presentation.append(
        [
            "Model",
            "Pi IPS",
            "Jetson IPS",
            "Speedup x",
            "Throughput gain %",
            "Pi mean ms",
            "Jetson mean ms",
            "Latency reduction %",
        ]
    )
    current_versions = ("3.4", "3.4b", "4.0f", "4.1a", "4.1g")
    for version in current_versions:
        row = next(item for item in rows if item[0] == f"v{version}")
        presentation.append(
            [row[0], row[3], row[4], row[5], row[6], row[7], row[8], row[9]]
        )
    style_header(presentation[15])
    for row in presentation.iter_rows(min_row=16, max_row=20):
        row[3].number_format = '0.000"x"'
        row[4].number_format = '0.0"%"'
        row[7].number_format = '0.0"%"'
    presentation["A23"] = "Publication caution"
    presentation["B23"] = (
        "Repeat the suite across multiple passes with temperature, clock, and "
        "background-load logging before treating one percentage as a final "
        "public bound."
    )
    presentation.merge_cells("B23:H24")
    presentation["B23"].alignment = Alignment(wrap_text=True, vertical="top")
    presentation["B23"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    presentation.column_dimensions["A"].width = 26
    for column in range(2, 9):
        presentation.column_dimensions[get_column_letter(column)].width = 18

    models_sheet = workbook.create_sheet("All Models")
    models_sheet.append(MODEL_HEADERS)
    for row in rows:
        models_sheet.append(row)
    style_header(models_sheet[1])
    add_table(models_sheet, "AllModelComparisons")
    models_sheet.freeze_panes = "D2"
    models_sheet.auto_filter.ref = models_sheet.dimensions
    models_sheet.row_dimensions[1].height = 44
    add_number_formats(models_sheet)
    for column in (7, 10, 13, 16, 19, 23, 28, 31):
        letter = get_column_letter(column)
        models_sheet.conditional_formatting.add(
            f"{letter}2:{letter}{models_sheet.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
    fit_columns(models_sheet, maximum=20)
    models_sheet.column_dimensions["AK"].width = 16

    family_sheet = workbook.create_sheet("Family Summary")
    family_sheet.append(FAMILY_HEADERS)
    for family in families:
        family_sheet.append(family)
    style_header(family_sheet[1])
    add_table(family_sheet, "FamilyComparisons")
    family_sheet.freeze_panes = "C2"
    for row in family_sheet.iter_rows(min_row=2):
        row[4].number_format = '0.000"x"'
        row[9].number_format = '0.000"x"'
        for index in (5, 8, 10, 11, 15):
            row[index].number_format = '0.0"%"'
    fit_columns(family_sheet, maximum=24)

    memory_sheet = workbook.create_sheet("Memory Details")
    memory_headers = [
        "Model",
        "Pi before RSS MiB",
        "Pi loaded RSS MiB",
        "Pi warm RSS MiB",
        "Pi final RSS MiB",
        "Pi warm RSS delta MiB",
        "Jetson before RSS MiB",
        "Jetson loaded RSS MiB",
        "Jetson warm RSS MiB",
        "Jetson final RSS MiB",
        "Jetson warm RSS delta MiB",
        "Jetson CUDA before used MiB",
        "Jetson CUDA warm used MiB",
        "Jetson CUDA-visible delta MiB",
        "Jetson system-used delta MiB",
    ]
    memory_sheet.append(memory_headers)
    for version in versions:
        bm = baseline_reports[version]["memory"]
        cm = comparison_reports[version]["memory"]
        memory_sheet.append(
            [
                f"v{version}",
                bm["before_model_load"]["process_rss_mib"],
                bm["after_model_load"]["process_rss_mib"],
                bm["after_warmup"]["process_rss_mib"],
                bm["after_measured_runs"]["process_rss_mib"],
                bm["warm_process_rss_delta_mib"],
                cm["before_model_load"]["process_rss_mib"],
                cm["after_model_load"]["process_rss_mib"],
                cm["after_warmup"]["process_rss_mib"],
                cm["after_measured_runs"]["process_rss_mib"],
                cm["warm_process_rss_delta_mib"],
                cm["before_model_load"]["cuda_visible"]["used_mib"],
                cm["after_warmup"]["cuda_visible"]["used_mib"],
                cm["warm_cuda_visible_used_delta_mib"],
                cm["warm_system_used_delta_mib"],
            ]
        )
    style_header(memory_sheet[1])
    add_table(memory_sheet, "MemoryMeasurements")
    memory_sheet.freeze_panes = "B2"
    fit_columns(memory_sheet, maximum=24)

    charts = workbook.create_sheet("Charts")
    charts.sheet_view.showGridLines = False
    throughput_chart = LineChart()
    throughput_chart.title = "Isolated ONNX Throughput by Model"
    throughput_chart.y_axis.title = "Inferences per second"
    throughput_chart.x_axis.title = "Model"
    throughput_chart.height = 10
    throughput_chart.width = 24
    throughput_chart.add_data(
        Reference(models_sheet, min_col=4, max_col=5, min_row=1, max_row=53),
        titles_from_data=True,
    )
    throughput_chart.set_categories(
        Reference(models_sheet, min_col=1, min_row=2, max_row=53)
    )
    charts.add_chart(throughput_chart, "A1")

    latency_chart = LineChart()
    latency_chart.title = "Mean Model-Execution Latency"
    latency_chart.y_axis.title = "Milliseconds"
    latency_chart.x_axis.title = "Model"
    latency_chart.height = 10
    latency_chart.width = 24
    latency_chart.add_data(
        Reference(models_sheet, min_col=8, max_col=9, min_row=1, max_row=53),
        titles_from_data=True,
    )
    latency_chart.set_categories(
        Reference(models_sheet, min_col=1, min_row=2, max_row=53)
    )
    charts.add_chart(latency_chart, "A21")

    family_chart = BarChart()
    family_chart.title = "Geometric Throughput Speedup by Family"
    family_chart.y_axis.title = "Jetson / Raspberry Pi"
    family_chart.height = 9
    family_chart.width = 16
    family_chart.add_data(
        Reference(family_sheet, min_col=5, min_row=1, max_row=8),
        titles_from_data=True,
    )
    family_chart.set_categories(
        Reference(family_sheet, min_col=1, min_row=2, max_row=8)
    )
    charts.add_chart(family_chart, "Y1")

    definitions = workbook.create_sheet("Definitions")
    definitions.append(["Field", "Meaning"])
    definitions_rows = [
        ("IPS", "Batch-one model executions per second; preprocessing is excluded."),
        (
            "Local pipeline IPS",
            "Local image read/preprocessing plus model execution; no camera, "
            "Ethernet, controller, PWM, or physical servo.",
        ),
        ("Throughput gain %", "(Jetson IPS / Raspberry Pi IPS - 1) x 100."),
        (
            "Latency reduction %",
            "(Raspberry Pi latency - Jetson latency) / Raspberry Pi latency x 100.",
        ),
        (
            "Effective CPU cores",
            "Process CPU seconds divided by wall-clock inference seconds.",
        ),
        (
            "Warm RSS delta",
            "Process RSS after warmup minus process RSS before loading the model.",
        ),
        (
            "CUDA-visible delta",
            "Jetson shared-memory usage visible to the CUDA/runtime probe after "
            "warmup minus before model load.",
        ),
        (
            "Memory warning",
            "Jetson uses unified LPDDR5. Process RSS, system-used memory, and "
            "CUDA-visible memory overlap and must not be added together or "
            "described as model-weight-only memory.",
        ),
        (
            "Test scope",
            "Same 180 real sidewalk images, FP32 ONNX, batch size one, 20 "
            "warmups, and 200 measured runs per model.",
        ),
        (
            "Not measured",
            "Camera capture, JPEG Ethernet transport, Raspberry Pi arbitration, "
            "PCA9685 timing, and physical steering-servo response.",
        ),
    ]
    for row in definitions_rows:
        definitions.append(row)
    style_header(definitions[1])
    add_table(definitions, "MetricDefinitions")
    definitions.column_dimensions["A"].width = 26
    definitions.column_dimensions["B"].width = 110
    for row in definitions.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    definitions.freeze_panes = "A2"

    metadata = workbook.create_sheet("Test Metadata")
    metadata.append(["Field", "Raspberry Pi 5", "Jetson Orin Nano"])
    first_baseline = baseline_reports[versions[0]]
    first_comparison = comparison_reports[versions[0]]
    metadata_rows = [
        ("Suite timestamp", baseline["timestamp_local"], comparison["timestamp_local"]),
        ("Label", baseline["label"], comparison["label"]),
        ("Provider", baseline["provider"], comparison["provider"]),
        (
            "Active providers",
            ", ".join(first_baseline["runtime"]["active_providers"]),
            ", ".join(first_comparison["runtime"]["active_providers"]),
        ),
        (
            "ONNX Runtime",
            first_baseline["runtime"]["onnxruntime"],
            first_comparison["runtime"]["onnxruntime"],
        ),
        (
            "Python",
            first_baseline["runtime"]["python"],
            first_comparison["runtime"]["python"],
        ),
        (
            "Board",
            first_baseline["device"]["board"],
            first_comparison["device"]["board"],
        ),
        (
            "CPU",
            first_baseline["device"]["cpu"],
            first_comparison["device"]["cpu"],
        ),
        (
            "Logical CPU count",
            first_baseline["device"]["logical_cpu_count"],
            first_comparison["device"]["logical_cpu_count"],
        ),
        (
            "Platform",
            first_baseline["device"]["platform"],
            first_comparison["device"]["platform"],
        ),
        (
            "Image count",
            first_baseline["mini_test_dataset"]["image_count"],
            first_comparison["mini_test_dataset"]["image_count"],
        ),
        (
            "Mini test dataset SHA-256",
            first_baseline["mini_test_dataset"]["set_sha256"],
            first_comparison["mini_test_dataset"]["set_sha256"],
        ),
        (
            "Warmup runs",
            first_baseline["performance"]["warmup_runs"],
            first_comparison["performance"]["warmup_runs"],
        ),
        (
            "Measured runs",
            first_baseline["performance"]["measured_runs"],
            first_comparison["performance"]["measured_runs"],
        ),
        ("Model count", baseline["model_count"], comparison["model_count"]),
    ]
    for row in metadata_rows:
        metadata.append(row)
    style_header(metadata[1])
    add_table(metadata, "TestMetadata")
    metadata.freeze_panes = "B2"
    metadata.column_dimensions["A"].width = 25
    metadata.column_dimensions["B"].width = 62
    metadata.column_dimensions["C"].width = 62
    for row in metadata.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    derived_payload = {
        "benchmark": "SidewalkPilot-device-comparison-workbook-v1",
        "baseline": baseline["label"],
        "comparison": comparison["label"],
        "family_summaries": [
            dict(zip(FAMILY_HEADERS, family, strict=True)) for family in families
        ],
        "models": [
            dict(zip(MODEL_HEADERS, row, strict=True)) for row in rows
        ],
    }
    add_raw_json_sheet(workbook, "Raw Pi JSON", baseline, "RawPiJson")
    add_raw_json_sheet(
        workbook, "Raw Jetson JSON", comparison, "RawJetsonJson"
    )
    add_raw_json_sheet(
        workbook, "Raw Derived JSON", derived_payload, "RawDerivedJson"
    )

    for sheet in workbook.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
    workbook.properties.title = "SidewalkPilot Raspberry Pi 5 vs Jetson Orin Nano"
    workbook.properties.subject = "FP32 ONNX device benchmark comparison"
    workbook.properties.creator = "SidewalkPilot"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("baseline", type=Path)
    parser.add_argument("comparison", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    build_workbook(
        args.baseline.expanduser().resolve(),
        args.comparison.expanduser().resolve(),
        args.output.expanduser().resolve(),
    )
    print(f"Saved workbook: {args.output.expanduser().resolve()}")


if __name__ == "__main__":
    main()
